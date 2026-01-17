import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

def func_invoice(name, invoice_no, current_date, utr, nin, week_ending, company_name, company_road, company_city, 
                company_postcode, site_mon, pay_mon, site_tues, pay_tues, site_wed, pay_wed, site_thurs, 
                pay_thurs, site_fri, pay_fri, site_sat, pay_sat, site_sun, pay_sun, bank_name, sort_code, account_no, expenses):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # --- 1. STYLING SETUP ---
    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid") # Dark Grey
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    currency_format = '£#,##0.00'

    # Set Column Widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    # --- 2. HEADER SECTION ---
    ws.merge_cells("A1:E2")
    ws["A1"] = "SUBCONTRACTOR INVOICE"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A1"].alignment = align_center

    # Invoice Details (Top Right)
    ws["D4"] = "Invoice No:"
    ws["E4"] = invoice_no
    ws["D5"] = "Date:"
    ws["E5"] = current_date
    ws["D6"] = "Week Ending:"
    ws["E6"] = week_ending

    for row in range(4, 7):
        ws[f"D{row}"].font = bold_font
        ws[f"E{row}"].alignment = align_right

    # --- 3. FROM / TO SECTIONS ---
    # From (Left)
    ws["A4"] = "FROM:"
    ws["A4"].font = bold_font
    ws["A5"] = name
    ws["A6"] = f"UTR: {utr}"
    ws["A7"] = f"NIN: {nin}"

    # To (Left - Lower)
    ws["A10"] = "BILL TO:"
    ws["A10"].font = bold_font
    ws["A11"] = company_name
    ws["A12"] = company_road
    ws["A13"] = f"{company_city}, {company_postcode}"

    # --- 4. JOB TABLE ---
    # Table Headers
    ws["A16"] = "Day"
    ws["B16"] = "Description / Site"
    ws["C16"] = "Rate (£)"
    
    for cell in ["A16", "B16", "C16"]:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = align_center
        ws[cell].border = thin_border

    # Data Rows
    days_data = [
        ("Monday", site_mon, pay_mon),
        ("Tuesday", site_tues, pay_tues),
        ("Wednesday", site_wed, pay_wed),
        ("Thursday", site_thurs, pay_thurs),
        ("Friday", site_fri, pay_fri),
        ("Saturday", site_sat, pay_sat),
        ("Sunday", site_sun, pay_sun)
    ]

    current_row = 17
    total_gross = 0

    for day, site, pay in days_data:
        try:
            pay_val = float(pay)
        except ValueError:
            pay_val = 0.0
        
        # Only add row if there is a site or pay
        if site or pay_val > 0:
            ws.cell(row=current_row, column=1, value=day).border = thin_border
            ws.cell(row=current_row, column=2, value=site).border = thin_border
            
            c_cell = ws.cell(row=current_row, column=3, value=pay_val)
            c_cell.border = thin_border
            c_cell.number_format = currency_format
            
            total_gross += pay_val
            current_row += 1

    # Ensure at least some empty rows if no data
    if current_row == 17: 
        current_row = 24

    # --- 5. CALCULATIONS & TOTALS ---
    try:
        expenses_val = float(expenses)
    except ValueError:
        expenses_val = 0.0

    tax_val = total_gross * 0.20
    net_labour = total_gross - tax_val
    total_payable = net_labour + expenses_val

    # Totals Block (Right Aligned under table)
    r = current_row + 2
    
    labels = [("Gross Pay:", total_gross), ("CIS Tax (20%):", -tax_val), ("Expenses:", expenses_val)]
    
    for label, val in labels:
        ws[f"B{r}"] = label
        ws[f"B{r}"].alignment = align_right
        ws[f"B{r}"].font = bold_font
        
        ws[f"C{r}"] = val
        ws[f"C{r}"].number_format = currency_format
        ws[f"C{r}"].border = thin_border
        r += 1

    # Grand Total
    r += 1
    ws[f"B{r}"] = "TOTAL PAYABLE:"
    ws[f"B{r}"].alignment = align_right
    ws[f"B{r}"].font = Font(bold=True, size=12)
    
    ws[f"C{r}"] = total_payable
    ws[f"C{r}"].number_format = currency_format
    ws[f"C{r}"].font = Font(bold=True, size=12)
    ws[f"C{r}"].border = thin_border
    ws[f"C{r}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow highlight

    # --- 6. BANK DETAILS (Bottom Left) ---
    b_row = current_row + 2
    ws[f"A{b_row}"] = "PAYMENT DETAILS"
    ws[f"A{b_row}"].font = bold_font
    ws[f"A{b_row+1}"] = f"Bank: {bank_name}"
    ws[f"A{b_row+2}"] = f"Sort Code: {sort_code}"
    ws[f"A{b_row+3}"] = f"Account No: {account_no}"

    # Save
    file_name = re.sub(r"\s+", "_", name).lower() + "_invoice.xlsx"
    wb.save(file_name)
    return file_name
